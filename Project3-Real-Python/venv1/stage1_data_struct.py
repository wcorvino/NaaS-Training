from openpyxl import load_workbook
workbook = load_workbook(filename="/home/wcor/PycharmProjects/NaaS-Training/Project3-Real-Python/venv1/sample.xlsx")
sheet = workbook.active
"""
 Re; Working with spreadsheet Tuples (immutable lists)
 convert tuples into data structures (key:value pairs - dictionary)
 create a dictionary where the key is product_id
 Logic - iterate thru rows, grab product_id, then 
 collect data in related columns in another dictionary.
 Create a dictionary of dictionaries, indexed on product_id
"""

# grab headers!!
for value in sheet.iter_rows(min_row=1,
                             max_row=1,
                             values_only=True):
    print(value)

# grab data
for value in sheet.iter_rows(min_row=2,
                             max_row=4,
                             max_col=7,
                             values_only=True):
    print(value)

