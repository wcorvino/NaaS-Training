import json
from openpyxl import load_workbook
"""
map header & data discovered in stage 1, into a dictionary
"""
workbook = load_workbook(filename='/home/wcor/PycharmProjects/NaaS-Training/Project3-Real-Python/venv1/sample.xlsx')
sheet = workbook.active

products = {}

for item in sheet.iter_rows(min_row=2, min_col=4, max_col=7,
                           values_only=True):
    """
    item columns ; 0 = product_id , 1 = parent, 2 = title, 3 = category
    """
    product_id = item[0]
    product_info = {
        "parent": item[1],
        "title": item[2],
        "category": item[3],
    }
    """
    create a dictionary for key:value pairs, dictionaries
    dict products, indexed on product_id, is eq to product struct (info on the product)
    """
    products[product_id] = product_info

print(json.dumps(products))
"""
    Sample output
     dict(products)  ______dict(product_info)
     |              |
     v              v
    {"B00FALQ1ZC": {"parent": 937001370, "title": "Invicta Women's 15150 \"Angel\" 18k Yellow Gold Ion-Plated Stainless Steel and Brown Leather Watch", "category": "Watches"}}
    
    Call a value;
    products['B00FALQ1ZC']['parent']
    937001370
    
    call type
    type(products['B00FALQ1ZC']['parent'])
    <class 'int'>
    
"""
