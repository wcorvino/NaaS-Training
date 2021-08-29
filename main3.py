import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

#
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    # update column 5
    inventory_price = product_list.cell(product_row, 5)


    # ca
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        print("Added new supplier")
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculate  for inv less than 10
    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)

    # add value for inv price to spreadsheet , column 5
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)

product_list.cell(1, 5).value = ".py calculated total"
inv_file.save("inventory_rendered_by_python.xlsx")



